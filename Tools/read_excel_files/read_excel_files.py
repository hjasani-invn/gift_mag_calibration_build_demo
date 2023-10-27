# Python function to read an excel file

import openpyxl    #reading/writing Excel table
import statistics

def process_sheet(wb_obj, sheet, wb_out, sheet_out, cur_row, cur_column):
    print(sheet, end=',')
    sheet_obj = wb_obj[sheet]
    n = 4
    avail_data = []
    avg_data = []
    ttff_data = []
    rms_data = []
    max_data = []

    cell_name = sheet_obj.cell(row=n, column=2)
    cell_avail_data = sheet_obj.cell(row=n, column=4)
    cell_avg_data = sheet_obj.cell(row=n, column=5)
    cell_ttff_data = sheet_obj.cell(row=n, column=15)
    cell_rms_data = sheet_obj.cell(row=n, column=6)
    cell_max_data = sheet_obj.cell(row=n, column=7)

    count = 0
    while cell_name.value != "Average":
        avail_data.append(cell_avail_data.value)
        avg_data.append(cell_avg_data.value)
        ttff_data.append(cell_ttff_data.value)
        rms_data.append(cell_rms_data.value)
        max_data.append(cell_max_data.value)

        n = n + 1

        cell_avail_data = sheet_obj.cell(row=n, column=4)
        cell_avg_data = sheet_obj.cell(row=n, column=5)
        cell_ttff_data = sheet_obj.cell(row=n, column=15)
        cell_rms_data = sheet_obj.cell(row=n, column=6)
        cell_max_data = sheet_obj.cell(row=n, column=7)
        count = count + 1
        cell_name = sheet_obj.cell(row=n, column=2)

    print('%d %s %.2f' % (count, " , ", statistics.mean(avail_data) * 100), end=',')
    print('%.3f %s %.3f' % (statistics.mean(avg_data), " , ", statistics.median(avg_data)), end=',')
    print('%.3f %s %.3f' % (statistics.mean(ttff_data), " , ", statistics.median(ttff_data)), end=',')
    print('%.3f %s %.3f' % (statistics.mean(rms_data), " , ", statistics.median(rms_data)), end=',')
    print('%.3f %s %.3f' % (statistics.mean(max_data), " , ", statistics.median(max_data)))

    sheet_out.cell(row=cur_row+1, column=cur_column).value = sheet
    sheet_out.cell(row=cur_row+2, column=cur_column).value = count

    sheet_out.cell(row=cur_row+3, column=cur_column).number_format = '#,##0.00'
    sheet_out.cell(row=cur_row+3, column=cur_column).value = statistics.mean(avail_data) * 100

    sheet_out.cell(row=cur_row+4, column=cur_column).number_format = '#,##0.000'
    sheet_out.cell(row=cur_row+4, column=cur_column).value = statistics.mean(avg_data)

    sheet_out.cell(row=cur_row+5, column=cur_column).number_format = '#,##0.000'
    sheet_out.cell(row=cur_row+5, column=cur_column).value = statistics.median(avg_data)

    sheet_out.cell(row=cur_row+6, column=cur_column).number_format = '#,##0.000'
    sheet_out.cell(row=cur_row+6, column=cur_column).value = statistics.mean(ttff_data)

    sheet_out.cell(row=cur_row+7, column=cur_column).number_format = '#,##0.000'
    sheet_out.cell(row=cur_row+7, column=cur_column).value = statistics.median(ttff_data)

    sheet_out.cell(row=cur_row+8, column=cur_column).number_format = '#,##0.000'
    sheet_out.cell(row=cur_row+8, column=cur_column).value = statistics.mean(rms_data)

    sheet_out.cell(row=cur_row+9, column=cur_column).number_format = '#,##0.000'
    sheet_out.cell(row=cur_row+9, column=cur_column).value = statistics.median(rms_data)

    sheet_out.cell(row=cur_row+10, column=cur_column).number_format = '#,##0.000'
    sheet_out.cell(row=cur_row+10, column=cur_column).value = statistics.mean(max_data)

    sheet_out.cell(row=cur_row+11, column=cur_column).number_format = '#,##0.000'
    sheet_out.cell(row=cur_row+11, column=cur_column).value = statistics.median(max_data)


# Give the location of the file
def process_all_sheets(excel_file, wb_out , cur_row):

    print("")
    print(excel_file)
    print('%s %s %s %s %s %s %s %s %s %s %s %s %s %s %s %s %s %s %s %s %s' \
          % ("sheet name", " , ","count", " , ", "availability %", " , ", \
             "avg mean", " , " , "avg median", " , ", "ttff mean", " , " , "ttff median", \
             " , ", "rms mean", " , ", "rms median"," , ", "rms mean", " , ", "rms median"))

    sheet_out = wb_out.active
    cur_column = 1

    sheet_out.cell(row=cur_row, column=cur_column).value = excel_file

    sheet_out.cell(row=cur_row+1, column=cur_column).value = "sheet name"
    sheet_out.cell(row=cur_row+2, column=cur_column).value = "count"
    sheet_out.cell(row=cur_row+3, column=cur_column).value = "availability %"
    sheet_out.cell(row=cur_row+4, column=cur_column).value = "avg mean"
    sheet_out.cell(row=cur_row+5, column=cur_column).value = "avg median"
    sheet_out.cell(row=cur_row+6, column=cur_column).value = "ttff mean"
    sheet_out.cell(row=cur_row+7, column=cur_column).value = "ttff median"
    sheet_out.cell(row=cur_row+8, column=cur_column).value = "rms mean"
    sheet_out.cell(row=cur_row+9, column=cur_column).value = "rms median"
    sheet_out.cell(row=cur_row+10, column=cur_column).value = "max mean"
    sheet_out.cell(row=cur_row+11, column=cur_column).value = "max median"

    # To open the workbook
    # workbook object is created
    wb_obj = openpyxl.load_workbook(excel_file)

    cur_column = 3
    sheet_list =wb_obj.sheetnames
    for m in range(0, len(sheet_list)):
        if sheet_list[m].startswith("mixed_"):
            process_sheet(wb_obj, sheet_list[m], wb_out, sheet_out, cur_row, cur_column)
            cur_column = cur_column + 1
    for m in range(0, len(sheet_list)):
        if sheet_list[m].startswith("mag_"):
            process_sheet(wb_obj, sheet_list[m], wb_out, sheet_out, cur_row, cur_column)
            cur_column = cur_column + 1
    for m in range(0, len(sheet_list)):
        if not (sheet_list[m].startswith("mag_") or sheet_list[m].startswith("mixed_")):
            process_sheet(wb_obj, sheet_list[m], wb_out, sheet_out, cur_row, cur_column)
            cur_column = cur_column + 1

    cur_row = cur_row + 13
    return cur_row
