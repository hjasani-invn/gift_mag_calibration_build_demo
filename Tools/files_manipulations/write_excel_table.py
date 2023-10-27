# Python function to read an excel file

import openpyxl    #reading/writing Excel table

def write_to_excel(tracks, filename):
    wb_out = openpyxl.Workbook()
    sheet_out = wb_out.active
    sheet_out.title = 'multi floor results'
    cur_row = 4
    cur_column = 1

    max_row = write_sheet(tracks,  sheet_out, cur_row, cur_column)

    sheet_out.cell(row=cur_row - 3, column=cur_column).value = "dataset"
    sheet_out.cell(row=cur_row - 3, column=cur_column + 1).value = "positive"
    sheet_out.cell(row=cur_row - 3, column=cur_column + 2).value = "negative"
    sheet_out.cell(row=cur_row - 3, column=cur_column + 3).value = "flickering"

    sheet_out.cell(row=cur_row - 1, column=cur_column).value = "threshold"
    sheet_out.cell(row=cur_row - 1, column=cur_column + 1).value = 1  # threshold_B
    sheet_out.cell(row=cur_row - 1, column=cur_column + 2).value = 1  # threshold_B
    sheet_out.cell(row=cur_row - 1, column=cur_column + 3).value = 1  # threshold_B

    dB = (sheet_out.cell(row=cur_row, column=cur_column + 1).coordinate + ":" +
          sheet_out.cell(row=max_row, column=cur_column + 1).coordinate)
    dC = (sheet_out.cell(row=cur_row, column=cur_column + 2).coordinate + ":" +
          sheet_out.cell(row=max_row, column=cur_column + 2).coordinate)
    dD = (sheet_out.cell(row=cur_row, column=cur_column + 3).coordinate + ":" +
          sheet_out.cell(row=max_row, column=cur_column + 3).coordinate)

    # =COUNTIF(B4:B13,">"&B3)
    count_B = '=COUNTIF(' + dB + ',">="&' + sheet_out.cell(row=cur_row - 1, column=cur_column + 1).coordinate + ')'
    count_C = '=COUNTIF(' + dC + ',">="&' + sheet_out.cell(row=cur_row - 1, column=cur_column + 2).coordinate + ')'
    count_D = '=COUNTIF(' + dD + ',">="&' + sheet_out.cell(row=cur_row - 1, column=cur_column + 3).coordinate + ')'

    sheet_out.cell(row=cur_row - 2, column=cur_column).value = "count"
    sheet_out.cell(row=cur_row - 2, column=cur_column + 1).value = count_B
    sheet_out.cell(row=cur_row - 2, column=cur_column + 2).value = count_C
    sheet_out.cell(row=cur_row - 2, column=cur_column + 3).value = count_D

    wb_out.save(filename)


def write_sheet(tracks,  sheet_out, cur_row, cur_column):
    n = 0
    for track in tracks:
        sheet_out.cell(row=cur_row + n, column=cur_column).value = track[0]
        sheet_out.cell(row=cur_row + n, column=cur_column + 1).number_format = '#,##0.00'
        sheet_out.cell(row=cur_row + n, column=cur_column + 1).value = float(track[3])
        sheet_out.cell(row=cur_row + n, column=cur_column + 2).number_format = '#,##0.00'
        sheet_out.cell(row=cur_row + n, column=cur_column + 2).value = float(track[4])
        sheet_out.cell(row=cur_row + n, column=cur_column + 3).number_format = '#,##0.00'
        sheet_out.cell(row=cur_row + n, column=cur_column + 3).value = float(track[5])

        n += 1

    return (cur_row + n - 1)
